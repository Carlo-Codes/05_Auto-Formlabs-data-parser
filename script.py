import openpyxl
import csv
from pathlib import Path
import tkinter as tk
from tkinter.filedialog import askopenfilename 
import re
from datetime import datetime 
from openpyxl.styles import colors, PatternFill



COLUMNS_TO_EXTRACT_FROM_CSV = ["Print name", "Username", "Printer", "Status", "Success?", "Elapsed print time (ms)", "Layers", "Layer thickness (um)", "Volume (ml)", "Material", "Finish time", "Start time"]
EXCEL_COLUMN_MAP = {
    "JOB NUMBER" : 1,
    "Print name" : 2,
    "Username" : 3,
    "Printer" : 4,
    "Status" : 5,
    "Success?" : 6,
    "Start time" : 7,
    "Finish time" : 8,
    "Elapsed print time (ms)" : 9,
    "Layers" : 10,
    "Layer thickness (um)" : 11,
    "Volume (ml)" : 12,
    "Material" : 13,
}
STARTIME_COLUMN = "G"

VOLUME_FILL = PatternFill(start_color='8EA9DB', end_color="8EA9DB", fill_type='solid')

mainFile = ""
exportedDataFile = ""
dirName = Path.cwd()
dataToInsert = {}


def openMainFile():
    global dirName
    global mainFile
    filename = askopenfilename(filetypes=[("Excel Files", "*.xlsx")], initialdir= dirName)
    mainFile = filename
    tk.Label(window,text=filename).grid(column=1, row=0)

def openExportedDataFile():
    global dirName
    global exportedDataFile
    filename = askopenfilename(filetypes=[("CSV Files", "*.csv")], initialdir= dirName)
    exportedDataFile = filename
    tk.Label(window,text=filename).grid(column=1, row=1)

def ok():

    extractedData = extractCSVdata(exportedDataFile)
    processedData = AddJobNumberToData(extractedData)
    #insertDataIntoXlFile(mainFile, processedData)
    latestDateinData = findLatestRowInExistingData(mainFile)
    print(latestDateinData)
    filteredData = FilterOutOldDataFromNew(processedData, latestDateinData)
    insertDataIntoXlFile(mainFile,filteredData)

    


def extractCSVdata(file):
    dataToReturn = []
    
    with open(file,mode='r') as csv_file:
        line = 0
        csvReader = csv.DictReader(csv_file)
        for row in csvReader:
            dataToAdd = {}
            if line == 0:
                line += 1
            for cell in row:
                for columnName in COLUMNS_TO_EXTRACT_FROM_CSV:
                    if cell == columnName:
                         dataToAdd[cell] = row[cell]
            dataToReturn.append(dataToAdd)
    
    return dataToReturn
            
def generateJobNumberPrintName(name):
    results = re.findall("[SCTPOWEAGsctpoweag][-_ ][0-9][0-9][0-9][0-9]|[SCTPOWEAGsctpoweag][0-9][0-9][0-9][0-9]", name)
    jobcode = ", ".join(results)
    jobcode = jobcode.upper()
    if not results:
        jobcode = "Unknown"
    return jobcode

def AddJobNumberToData(data):
    returnData = []
    for row in data:
        newRow = row
        newRow["JOB NUMBER"] = generateJobNumberPrintName(newRow["Print name"])
        returnData.append(newRow)
    return returnData

def insertDataIntoXlFile(filename, data):
    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[0]

     
    for row in range(len(data)):
        inseretionRow = ws.max_row+1
        for col in EXCEL_COLUMN_MAP:
            cellRef = ws.cell(row=inseretionRow,column=EXCEL_COLUMN_MAP[col])
            valueToAdd = ""
            try:
                valueToAdd = float(data[row][col])
            except:
                valueToAdd =data[row][col]
            cellRef.value = valueToAdd
            if col == "Volume (ml)":
                cellRef.fill = VOLUME_FILL

            
    
    wb.save("3D Print data{}.xlsx".format(datetime.now().strftime("%Y%m%d%H%M%S")))

def findLatestRowInExistingData(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[0]
    datetoCheck = datetime.min
    for row in ws[STARTIME_COLUMN]:
        try:
            datetoCheckagaints = datetime.fromisoformat(row.value)
            if datetoCheckagaints.date() > datetoCheck.date():
                datetoCheck = datetoCheckagaints
        except:
            print(row.value)
    
    return datetoCheck

def FilterOutOldDataFromNew(data, lastdate):
    dataToReturn = []
    for row in data:
        rowDate = datetime.fromisoformat(row["Start time"])
        if rowDate.date() > lastdate.date():
            dataToReturn.append(row)
    return dataToReturn



window = tk.Tk()
window.grid()

tk.Button(window, text="Select Main File", command = openMainFile).grid(column=0, row=0)

tk.Button(window, text="Select Exported Data File", command = openExportedDataFile).grid(column=0, row=1)

tk.Button(window, text="Ok", command=ok).grid(column=0, row=2)
tk.Button(window, text="Quit", command=window.destroy).grid(column=1, row=2)

window.mainloop()